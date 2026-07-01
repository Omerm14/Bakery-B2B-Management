"""
Import historical Excel order data into Supabase.
Generates scripts/seed_data.sql for pasting into Supabase SQL Editor.
Usage: python scripts/import_excel.py
"""

import re, sys
from pathlib import Path
from datetime import datetime, date, timedelta
import openpyxl

UPLOADS_DIR = Path("/root/.claude/uploads/040e7575-cbc0-5cc4-82fe-f34117c9403a")
OUT = Path(__file__).parent / "seed_data.sql"

SKIP_PATTERNS = [
    r"כמויות", r"כפתורי", r"גיליון", r'לו"ז', r"ללא מיתוג", r"רשימת",
    r"Sheet\d*$", r"לקוח חדש", r"דוגמאות", r"^\.*$", r"^[…—_ס]",
    r"סופגניות", r"בראסרי", r"המבורגר",
]
NON_ITEM_ROWS = {
    "מאפים", "מתוקים", "קפואים", "שונות", "עוגות ועוגיות",
    "קפואים ושונות - קונדי", "תאריך",
}
# Map Hebrew day names to weekday offsets from Sunday
DAY_OFFSETS = {"ראשון": 0, "שני": 1, "שלישי": 2, "רביעי": 3, "חמישי": 4, "שישי": 5, "שבת": 6, "שבת ": 6}


def should_skip(name):
    for p in SKIP_PATTERNS:
        if re.search(p, name):
            return True
    return False


def normalize_customer(name):
    name = re.sub(r"\s*-?\s*תפריט חדש\s*", "", name)
    name = re.sub(r"\s*-\s*\d+%\s*הנחה.*", "", name)
    name = name.replace("(ליגורי)", "").replace("עדינה", "").strip()
    # Remove trailing dashes and whitespace
    name = name.strip(" -")
    return name


def parse_qty(v):
    if v is None:
        return 0
    try:
        f = float(v)
        return f if f > 0 else 0
    except (TypeError, ValueError):
        return 0


def esc(s):
    return s.replace("'", "''")


def get_sunday(d):
    """Return the Sunday on or before date d."""
    dow = d.weekday()  # Mon=0 … Sun=6
    return d - timedelta(days=(dow + 1) % 7)


def parse_date_value(v):
    """Try to parse a cell value as a date."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and "/" in v:
        try:
            parts = v.strip().split("/")
            d2, m2, y2 = int(parts[0]), int(parts[1]), int(parts[2])
            if y2 < 100:
                y2 += 2000
            return date(y2, m2, d2)
        except Exception:
            pass
    return None


def parse_week_label(label):
    """Parse '28.1-3.2' style label into (start_date, end_date) or None."""
    m = re.match(r"^(\d{1,2})\.(\d{1,2})-(\d{1,2})\.(\d{1,2})$", label.strip())
    if m:
        d1, m1, d2, m2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        # Guess year from context (assume 2025 for the 3e31138d file)
        # We'll return the parsed start date; caller provides the year
        return (d1, m1), (d2, m2)
    return None


def extract_dates_from_sheet(ws, week_start):
    """
    Return list of 7 dates (cols B-H) from a customer sheet.
    Tries row 1 first (datetime/string). Falls back to day-name mapping.
    """
    dates = []
    # Try row 1 dates
    for col in range(2, 9):
        d = parse_date_value(ws.cell(1, col).value)
        dates.append(d)

    if any(d is not None for d in dates):
        return dates

    # Fall back: row 1 or row 2 has day names — use week_start + offset
    for row in (1, 2):
        row_vals = [ws.cell(row, col).value for col in range(2, 9)]
        if any(v in DAY_OFFSETS for v in row_vals if isinstance(v, str)):
            result = []
            for v in row_vals:
                if isinstance(v, str) and v.strip() in DAY_OFFSETS:
                    result.append(week_start + timedelta(days=DAY_OFFSETS[v.strip()]))
                else:
                    result.append(None)
            return result

    return [None] * 7


def find_week_start_for_file(wb):
    """
    Find the week start date for a workbook.
    Strategy 1: find a customer sheet with dates in row 1.
    Strategy 2: find a sheet named like '28.1-3.2' and parse it.
    """
    # Strategy 1: look for customer sheet with row-1 dates
    for sname in wb.sheetnames:
        if should_skip(sname):
            continue
        ws = wb[sname]
        for col in range(2, 9):
            d = parse_date_value(ws.cell(1, col).value)
            if d:
                return get_sunday(d)

    # Strategy 2: sheet named with date range
    for sname in wb.sheetnames:
        parsed = parse_week_label(sname)
        if parsed:
            (d1, m1), _ = parsed
            # Need to guess the year — scan sheetnames for a year hint in the filename
            # We'll return a partial and let the caller provide year
            return parsed  # caller handles

    return None


def main():
    all_customers = set()
    all_items = set()
    all_weeks = {}  # iso_str → label
    order_lines = []  # (week_start_iso, customer_name, item_name, delivery_date_iso, qty)

    for xlsx_file in sorted(UPLOADS_DIR.glob("*.xlsx")):
        print(f"\nReading {xlsx_file.name} ...")
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)

        # Determine week start for this file
        week_start = find_week_start_for_file(wb)

        if week_start is None:
            print("  → Could not determine week dates, skipping")
            continue

        # Handle the parsed-tuple case (date-range sheet name, no year from cells)
        year_override = None
        if isinstance(week_start, tuple):
            # Parse the tuple: ((d1, m1), (d2, m2))
            (d1, m1), _ = week_start
            # Guess year from filename
            fname = xlsx_file.name
            year_m = re.search(r"20(\d{2})", fname)
            if year_m:
                year = 2000 + int(year_m.group(1))
            else:
                # Try: 3e31138d looks like Jan 2025 based on content
                year = 2025
            try:
                start_d = date(year, m1, d1)
            except ValueError:
                print(f"  → Bad date {d1}/{m1}/{year}, skipping")
                continue
            week_start = get_sunday(start_d)
            year_override = year

        ws_str = week_start.isoformat()
        label = f"שבוע {week_start.day:02d}/{week_start.month:02d}"
        all_weeks[ws_str] = label
        print(f"  Week start: {ws_str} ({label})")

        for sheet_name in wb.sheetnames:
            if should_skip(sheet_name):
                continue
            ws = wb[sheet_name]
            customer_name = normalize_customer(sheet_name)
            if not customer_name or len(customer_name) < 2:
                continue

            dates = extract_dates_from_sheet(ws, week_start)
            if not any(d is not None for d in dates):
                print(f"  → No dates for sheet '{sheet_name}', skipping")
                continue

            # Determine item start row: skip header rows (row 1 = dates, row 2 = day names or category)
            item_start = 3
            r2_col1 = ws.cell(2, 1).value
            if r2_col1 and isinstance(r2_col1, str) and r2_col1 not in NON_ITEM_ROWS:
                # row 2 col 1 has a customer name or category — items start at row 2
                item_start = 2

            all_customers.add(customer_name)
            sheet_lines = 0

            for row in range(item_start, 200):
                item_name = ws.cell(row, 1).value
                if item_name is None:
                    continue
                if not isinstance(item_name, str):
                    continue
                item_name = item_name.strip()
                if not item_name or len(item_name) < 2:
                    continue
                if item_name.startswith("סה") or item_name in NON_ITEM_ROWS:
                    continue

                for col_idx, delivery_date in enumerate(dates):
                    if delivery_date is None:
                        continue
                    qty = parse_qty(ws.cell(row, col_idx + 2).value)
                    if qty > 0:
                        all_items.add(item_name)
                        order_lines.append((ws_str, customer_name, item_name, delivery_date.isoformat(), qty))
                        sheet_lines += 1

            if sheet_lines:
                print(f"  {customer_name}: {sheet_lines} lines")

    print(f"\n{'='*50}")
    print(f"Totals: {len(all_customers)} customers, {len(all_items)} items, {len(all_weeks)} weeks, {len(order_lines)} order lines")

    # ── Generate SQL ──
    lines = []
    lines.append("-- ============================================================")
    lines.append("-- Bakery B2B — Historical data seed")
    lines.append("-- Paste into Supabase SQL Editor and click RUN")
    lines.append("-- ============================================================\n")

    # Customers (need unique constraint on name — add it if missing)
    lines.append("-- Add unique constraint on customers.name if not exists")
    lines.append("DO $$ BEGIN")
    lines.append("  IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'customers_name_key') THEN")
    lines.append("    ALTER TABLE customers ADD CONSTRAINT customers_name_key UNIQUE (name);")
    lines.append("  END IF;")
    lines.append("END $$;\n")

    lines.append("-- Add unique constraint on menu_items.name_he if not exists")
    lines.append("DO $$ BEGIN")
    lines.append("  IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'menu_items_name_he_key') THEN")
    lines.append("    ALTER TABLE menu_items ADD CONSTRAINT menu_items_name_he_key UNIQUE (name_he);")
    lines.append("  END IF;")
    lines.append("END $$;\n")

    # Customers
    lines.append("-- CUSTOMERS")
    lines.append("INSERT INTO customers (name, active) VALUES")
    cust_list = sorted(all_customers)
    for i, c in enumerate(cust_list):
        comma = "" if i == len(cust_list) - 1 else ","
        lines.append(f"  ('{esc(c)}', true){comma}")
    lines.append("ON CONFLICT (name) DO NOTHING;\n")

    # Menu items
    lines.append("-- MENU ITEMS")
    lines.append("INSERT INTO menu_items (name_he, unit, active) VALUES")
    item_list = sorted(all_items)
    for i, it in enumerate(item_list):
        comma = "" if i == len(item_list) - 1 else ","
        lines.append(f"  ('{esc(it)}', 'יח׳', true){comma}")
    lines.append("ON CONFLICT (name_he) DO NOTHING;\n")

    # Weeks
    lines.append("-- WEEKS")
    lines.append("INSERT INTO weeks (start_date, label) VALUES")
    week_list = sorted(all_weeks.keys())
    for i, ws_str in enumerate(week_list):
        comma = "" if i == len(week_list) - 1 else ","
        lines.append(f"  ('{ws_str}', '{esc(all_weeks[ws_str])}'){comma}")
    lines.append("ON CONFLICT (start_date) DO NOTHING;\n")

    # Order lines
    lines.append("-- ORDER LINES")
    lines.append("INSERT INTO order_lines (week_id, customer_id, menu_item_id, delivery_date, quantity, source, status)")
    lines.append("SELECT")
    lines.append("  (SELECT id FROM weeks WHERE start_date = r.week_start::date),")
    lines.append("  (SELECT id FROM customers WHERE name = r.cname),")
    lines.append("  (SELECT id FROM menu_items WHERE name_he = r.iname),")
    lines.append("  r.delivery_date::date,")
    lines.append("  r.qty::numeric,")
    lines.append("  'manual',")
    lines.append("  'ok'")
    lines.append("FROM (VALUES")
    for i, (ws_str, cname, iname, ddate, qty) in enumerate(order_lines):
        comma = "" if i == len(order_lines) - 1 else ","
        lines.append(f"  ('{ws_str}', '{esc(cname)}', '{esc(iname)}', '{ddate}', {qty}){comma}")
    lines.append(") AS r(week_start, cname, iname, delivery_date, qty)")
    lines.append("WHERE")
    lines.append("  (SELECT id FROM weeks WHERE start_date = r.week_start::date) IS NOT NULL")
    lines.append("  AND (SELECT id FROM customers WHERE name = r.cname) IS NOT NULL")
    lines.append("  AND (SELECT id FROM menu_items WHERE name_he = r.iname) IS NOT NULL")
    lines.append("ON CONFLICT (week_id, customer_id, menu_item_id, delivery_date)")
    lines.append("DO UPDATE SET quantity = EXCLUDED.quantity;\n")

    lines.append(f"-- Done! {len(order_lines)} order lines, {len(all_customers)} customers, {len(all_items)} items, {len(all_weeks)} weeks")

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nSQL written to: {OUT}")
    print(f"File size: {OUT.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
